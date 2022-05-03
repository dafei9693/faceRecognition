from openpyxl import Workbook,load_workbook
import time
import os

def write(face_names,wb):
    '''day=time.strftime("%Y-%m-%d %H:%M:%S",time.localtime())[:10]
    try:
        wb=Workbook()
        wb.create_sheet("enter")
        wb.save("E:/Work/"+day+".xlsx")
    except PermissionError:
        wb=load_workbook("E:/Work/"+day+".xlsx")'''
    day=time.strftime("%Y-%m-%d %H:%M:%S",time.localtime())[:10]
    ws=wb["enter"]
    for face in face_names:
        if face[0] is not "Unknown":
            if_write = re_enert(face[0], ws.max_row, ws)
            if if_write:
                ws.append([face[0],time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())[11:]])
                break
            else:
                break
                '''if ws["A"+str(num)].value is None:
                    print("1")
                    if_write=re_enert(face[0], num, ws)
                    print(if_write)
                    if if_write== True:
                        ws["A" + str(num)].value = face[0]
                        ws["B" + str(num)].value = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())[11:]
                        print("writed")
                        break
                    if if_write == False:
                        break
                num += 1'''

    wb.save("E:/Work/"+day+".xlsx")
    return None

def open():
    day = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())[:10]
    if os.path.exists("E:/Work/" + day + ".xlsx"):
        wb = load_workbook("E:/Work/" + day + ".xlsx")
    else:
        wb = Workbook()
        wb.create_sheet("enter")
        wb.save("E:/Work/" + day + ".xlsx")

    '''try:
        wb = Workbook()
        wb.create_sheet("enter")
        wb.save("E:/Work/" + day + ".xlsx")
    except PermissionError:
        wb = load_workbook("E:/Work/" + day + ".xlsx")'''
    return wb

def re_enert(name,num,ws):
    now=time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())[11:]
    now=now.replace(":",'')
    for i in range(-num,0):
        Name=ws["A"+str(-i)].value
        Time=ws["B"+str(-i)].value
        try:
            Time=Time.replace(":",'')
        except AttributeError:
            break
        if now[:2] == Time[:2] :
            if int(now[2:4]) - int(Time[2:4]) > 2:
                break
        elif int(now[:2])>int(Time[:2]) :
            if int(now[2:4])+60 - int(Time[2:4]) >2:
                break
        if name == Name :
            return False

    return True
