import time
from tkinter import *
from tkinter import ttk
from openpyxl import load_workbook
import cv2
from PIL import Image,ImageTk
from excell_work import write,open
from face_recognize import face_recognition,face_rec


def get_info(names):
    if names == None:
        return None
    unknown=0
    for i in range(len(names)):
        if names[i] == "Unknown":
            unknown+=1
        if unknown == len(names):
            return None
    ws=load_workbook("E:/Work/信息.xlsx")["Sheet1"]
    info=[]
    for i in names:
        indivisual=[]
        num=1
        if i is not "Unknown":
            while True:
                num += 1
                if int(ws["B"+str(num)].value) == int(i):
                    indivisual.append(ws["A"+str(num)].value)
                    indivisual.append(ws["B"+str(num)].value)
                    info.append(indivisual)
                    break
    return info



def video():
    def back():
        win.destroy()

    wb=open()
    dududu = face_recognition()
    window_width = 1280
    window_height = 720
    image_width = int(window_width * 0.75)
    image_height = int(window_height)
    imagepos_x = 0
    imagepos_y = 0
    #butpos_x = 450
    #butpos_y = 450
    vc1 = cv2.VideoCapture(0)

    win = Toplevel()
    win.title("视频")
    win.geometry(str(window_width) + 'x' + str(window_height))
    canvas1 = Canvas(win, bg='white', width=image_width, height=image_height)
    canvas1.place(x=imagepos_x, y=imagepos_y)
    button1 = Button(win, text="退出", command=back)
    button1.place(x=image_width + 110, y=image_height - 50, width=100)
    lab1=Label(win,text="识别信息")
    lab1.place(x=image_width + 140, y=35)

    def tkImage(vc):
        ref, frame = vc.read()
        frame, face_names = dududu.recognize(frame)
        cvimage = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        pilImage = Image.fromarray(cvimage)
        pilImage = pilImage.resize((image_width, image_height), Image.ANTIALIAS)
        tkImage = ImageTk.PhotoImage(image=pilImage)
        return tkImage,face_names

    def video_loop():
        try:
            informations = StringVar()
            informations .set("")
            while True:
                text = ""
                time0 = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
                label = Label(win, text=time0)
                label.place(x=image_width + 55, y=10, width=200)
                picture1,face_names = tkImage(vc1)
                infos=get_info(face_names)
                if infos is not None:
                    for i in range(len(infos)):
                        text=text+"      姓名: "+infos[i][0]+"    学号: "+str(infos[i][1])+"\n\n"
                        informations.set(text)
                    write(infos,wb)
                if infos is None:
                    informations.set("")
                lab2 = Label(win, textvariable=informations)
                lab2.place(x=image_width + 27, y=60)
                canvas1.create_image(0,0, anchor='nw', image=picture1)
                win.update_idletasks()
                win.update()
        except:
            pass

    video_loop()
    win.mainloop()
    vc1.release()
    cv2.destroyAllWindows()

#video()
